"""MIS report — a multi-sheet Excel workbook summarising the whole ATS.

Sheet 1 is a KPI summary; the rest are row-level detail per entity. Built with
openpyxl and returned as raw bytes for the download endpoint.
"""

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.modules.auth.models import User
from app.modules.candidates.models import Candidate
from app.modules.dashboard.analytics import AnalyticsService
from app.modules.dashboard.service import DashboardService
from app.modules.interviews.models import Interview
from app.modules.jobs.models import Job
from app.modules.offers.models import Offer

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _naive(v):
    """openpyxl can't write tz-aware datetimes — drop the tzinfo."""
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    return v


def _ev(v):
    """Enum -> its value; pass everything else through."""
    return v.value if hasattr(v, "value") else v


class MISReportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def build(self) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty sheet

        await self._summary_sheet(wb)
        await self._candidates_sheet(wb)
        await self._jobs_sheet(wb)
        await self._interviews_sheet(wb)
        await self._offers_sheet(wb)
        await self._consultants_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------ #
    def _write(self, ws, headers, rows):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for row in rows:
            ws.append([_naive(v) for v in row])
        # Auto-ish column widths.
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            longest = len(str(headers[col - 1]))
            for row in rows:
                longest = max(longest, len(str(row[col - 1]) if row[col - 1] is not None else ""))
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)
        ws.freeze_panes = "A2"

    async def _summary_sheet(self, wb):
        rec = await AnalyticsService(self.db).recruiting()
        attr = await AnalyticsService(self.db).attrition()
        total_jobs = await self.db.scalar(select(func.count()).select_from(Job))
        open_jobs = await self.db.scalar(
            select(func.count()).select_from(Job).where(Job.status == "open")
        )
        total_cand = await self.db.scalar(select(func.count()).select_from(Candidate))

        ws = wb.create_sheet("Summary")
        rows = [
            ("Generated at", datetime.utcnow()),
            ("", ""),
            ("Jobs — total", total_jobs or 0),
            ("Jobs — open", open_jobs or 0),
            ("Candidates — total", total_cand or 0),
            ("Overall conversion %", rec["funnel"]["overall_conversion"]),
            ("Avg time to hire (days)", rec["time_to_hire"]["avg_days"]),
            ("Offer acceptance rate %", rec["offer_acceptance"]["acceptance_rate"]),
            ("Interviews completed", rec["interview_outcomes"]["completed"]),
            ("", ""),
            ("Active headcount", attr["headcount"]),
            ("Employees left", attr["left"]),
            ("Turnover rate %", attr["turnover_rate"]),
            ("Avg tenure (years)", attr["avg_tenure_years"]),
        ]
        self._write(ws, ["Metric", "Value"], rows)

    async def _candidates_sheet(self, wb):
        stmt = (
            select(
                Candidate.id,
                Candidate.full_name,
                Candidate.email,
                Candidate.phone,
                Job.title,
                Candidate.stage,
                Candidate.source,
                Candidate.priority,
                Candidate.ai_score,
                Candidate.created_at,
            )
            .join(Job, Candidate.job_id == Job.id, isouter=True)
            .order_by(Candidate.id.desc())
        )
        rows = [
            (r[0], r[1], r[2], r[3], r[4], _ev(r[5]), _ev(r[6]), _ev(r[7]), r[8], r[9])
            for r in (await self.db.execute(stmt)).all()
        ]
        ws = wb.create_sheet("Candidates")
        self._write(
            ws,
            ["ID", "Name", "Email", "Phone", "Job", "Stage", "Source", "Priority", "AI score", "Created"],
            rows,
        )

    async def _jobs_sheet(self, wb):
        consultant = aliased(User)
        count_sub = (
            select(Candidate.job_id, func.count(Candidate.id).label("n"))
            .group_by(Candidate.job_id)
            .subquery()
        )
        stmt = (
            select(
                Job.id,
                Job.title,
                Job.department,
                Job.location,
                Job.status,
                Job.positions,
                Job.priority,
                consultant.full_name,
                func.coalesce(count_sub.c.n, 0),
                Job.created_at,
            )
            .join(consultant, Job.assigned_consultant_id == consultant.id, isouter=True)
            .join(count_sub, count_sub.c.job_id == Job.id, isouter=True)
            .order_by(Job.id.desc())
        )
        rows = [
            (r[0], r[1], r[2], r[3], _ev(r[4]), r[5], _ev(r[6]), r[7], r[8], r[9])
            for r in (await self.db.execute(stmt)).all()
        ]
        ws = wb.create_sheet("Jobs")
        self._write(
            ws,
            ["ID", "Title", "Department", "Location", "Status", "Openings", "Priority", "Consultant", "Candidates", "Created"],
            rows,
        )

    async def _interviews_sheet(self, wb):
        manager = aliased(User)
        stmt = (
            select(
                Interview.id,
                Candidate.full_name,
                Job.title,
                Interview.mode,
                Interview.scheduled_at,
                Interview.status,
                Interview.outcome,
                manager.full_name,
            )
            .join(Candidate, Interview.candidate_id == Candidate.id, isouter=True)
            .join(Job, Candidate.job_id == Job.id, isouter=True)
            .join(manager, Interview.hiring_manager_id == manager.id, isouter=True)
            .order_by(Interview.scheduled_at.desc())
        )
        rows = [
            (r[0], r[1], r[2], _ev(r[3]), r[4], _ev(r[5]), _ev(r[6]), r[7])
            for r in (await self.db.execute(stmt)).all()
        ]
        ws = wb.create_sheet("Interviews")
        self._write(
            ws,
            ["ID", "Candidate", "Job", "Mode", "Scheduled", "Status", "Outcome", "Hiring manager"],
            rows,
        )

    async def _offers_sheet(self, wb):
        stmt = (
            select(
                Offer.id,
                Candidate.full_name,
                Offer.title,
                Offer.ctc,
                Offer.status,
                Offer.start_date,
                Offer.expiry_date,
                Offer.created_at,
            )
            .join(Candidate, Offer.candidate_id == Candidate.id, isouter=True)
            .order_by(Offer.id.desc())
        )
        rows = [
            (r[0], r[1], r[2], r[3], _ev(r[4]), r[5], r[6], r[7])
            for r in (await self.db.execute(stmt)).all()
        ]
        ws = wb.create_sheet("Offers")
        self._write(
            ws,
            ["ID", "Candidate", "Title", "CTC", "Status", "Start date", "Expiry", "Created"],
            rows,
        )

    async def _consultants_sheet(self, wb):
        data = await DashboardService(self.db).consultant_breakdown()
        rows = [
            (c["name"], c["total_jobs"], c["total_candidates"], c["submitted"])
            for c in data
        ]
        ws = wb.create_sheet("Consultant activity")
        self._write(ws, ["Consultant", "Jobs", "Candidates", "Submitted"], rows)
